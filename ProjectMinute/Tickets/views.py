from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
import openpyxl
from django.core.paginator import Paginator
from django.db import transaction
from .models import StateMaster, UserMaster, TaskMaster

def parser(request):
    if request.method == "GET":
        workbook = openpyxl.load_workbook("path_to_states.xlsx")
        sheet = workbook.active
        for state_name, state_code in sheet.iter_rows(min_row=2, values_only=True):
            StateMaster.objects.get_or_create(state_name=state_name, state_code=state_code)
        return HttpResponse("Data updated")

def userMaster(request):
    if request.method == "GET":
        workbook = openpyxl.load_workbook("path_to_userdata.xlsx")
        sheet = workbook.active
        states = StateMaster.objects.all()
        state_length = len(states)
        temp = 0
        for name, email, password, phone, city, gender, joining_date in sheet.iter_rows(min_row=2, values_only=True):
            UserMaster.objects.create(name=name, email=email, password=password, phone=phone,
                                      city=city, gender=gender, joining_date=joining_date, state=states[temp])
            temp = (temp + 1) % state_length
        return HttpResponse("Data created")

def bulkdata(request):
    if request.method == "GET":
        workbook = openpyxl.load_workbook("path_to_bulkdata.xlsx")
        sheet = workbook.active
        states = StateMaster.objects.all()
        state_length = len(states)
        temp = 0
        user_data = []
        for name, email, password, phone, city, gender, joining_date in sheet.iter_rows(min_row=2, values_only=True):
            if not UserMaster.objects.filter(name=name, email=email, phone=phone, joining_date=joining_date):
                user_data.append(UserMaster(name=name, email=email, password=password, phone=phone,
                                            city=city, gender=gender, joining_date=joining_date, state=states[temp]))
            temp = (temp + 1) % state_length
        with transaction.atomic():
            UserMaster.objects.bulk_create(user_data)
        return HttpResponse("Inserted")

def AssignTasks(request):
    if request.method == "POST":
        task_id = request.POST.get('taskid')
        title = request.POST.get('title')
        description = request.POST.get('description')
        assign_date = request.POST.get('assign_date')
        state = request.POST.get('state')
        assign_to = request.POST.get('assign_to')
        if UserMaster.objects.filter(id=assign_to).exists() and StateMaster.objects.filter(state_name=state).exists():
            TaskMaster.objects.create(task_id=task_id, title=title, description=description,
                                      assign_date=assign_date, state=StateMaster.objects.get(state_name=state),
                                      user_id=UserMaster.objects.get(id=assign_to))
        else:
            return HttpResponse("Something went wrong")
    add_state = StateMaster.objects.all()
    return render(request, "AssignTasks.html", {'states': add_state})

def homepage(request):
    states = StateMaster.objects.all()
    return render(request, 'index.html', {'states': states})

def aboutpage(request):
    return render(request, 'about.html')

def ViewTasks(request):
    items_per_page = 5
    page_number = request.GET.get('page', 1)
    search_query = request.GET.get('search')
    tasks = TaskMaster.objects.filter(title__icontains=search_query).order_by('id') if search_query else TaskMaster.objects.all().order_by('id')
    paginator = Paginator(tasks, items_per_page)
    paginated_data = paginator.get_page(page_number)
    total_pages = paginator.num_pages
    data = {'data': paginated_data, 'lastpage': total_pages, 'pageNumbers': [n+1 for n in range(total_pages)], 'search_query': search_query}
    return render(request, 'view.html', data)

def adduser(request):
    states = StateMaster.objects.all()
    if request.method == "POST":
        name = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phone = request.POST.get('phone')
        city = request.POST.get('city')
        gender = request.POST.get('gender')
        joining_date = request.POST.get('joining_date')
        state = request.POST.get('state')
        if StateMaster.objects.filter(state_name=state).exists():
            UserMaster.objects.create(name=name, email=email, password=password, phone=phone,
                                       city=city, gender=gender, joining_date=joining_date, state=StateMaster.objects.get(state_name=state))
            return render(request, 'adduser.html', {'states': states, 'added': True})
    return render(request, 'adduser.html', {'states': states})

def addmultipleusers(request):
    if request.method == "POST":
        file = request.FILES['excelFile']
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        states = StateMaster.objects.all()
        state_length = len(states)
        user_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, email, password, phone, city, gender, joining_date = row
            if not UserMaster.objects.filter(name=name, email=email, phone=phone, joining_date=joining_date).exists():
                user_data.append(UserMaster(name=name, email=email, password=password, phone=phone,
                                            city=city, gender=gender, joining_date=joining_date, state=states[user_data]))
            state = (state + 1) % state_length
        with transaction.atomic():
            UserMaster.objects.bulk_create(user_data)
        return render(request, 'addmultipleusers.html', {'added': True})
    return render(request, 'addmultipleusers.html')

def DeleteTask(request, id):
    task = get_object_or_404(TaskMaster, id=id)
    task.delete()
    return redirect('view-tasks')

def update(request, id):
    task = get_object_or_404(TaskMaster, id=id)
    if request.method == "POST":
        task_id = request.POST.get('taskid')
        title = request.POST.get('title')
        description = request.POST.get('description')
        assign_date = request.POST.get('assign_date')
        if assign_date:
            task.assign_date = assign_date
        task.task_id = task_id
        task.title = title
        task.description = description
        task.save()
        return redirect('view-tasks')
    return render(request, 'update.html', {'task': task})

from .serializers import StateSerializer
def add_state(request):
    if request.method == 'POST':
        state = StateSerializer(request.POST)
        if state.is_valid():
            state.save()
            return redirect('homepage') 
    else:
        state = StateSerializer()
    return render(request, 'add_state.html', {'state': state})







